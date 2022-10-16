#!/usr/bin/python3

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(filename = 'student.xlsx')
dest_filename = 'student.xlsx'
ws = wb.active
ws.title = "Sheet1"

dic = dict()
max_st = ws.max_row

for i in range(2, max_st + 1):
    midterm = ws.cell(row = i, column = 3).value
    final = ws.cell(row = i, column = 4).value
    hw = ws.cell(row = i, column = 5).value
    attendance = ws.cell(row = i, column = 6).value
    result = float(midterm*0.3 + final*0.35 + hw*0.34 + attendance*1)
    dic[i] = result
    ws.cell(row = i, column = 7, value = result)
result_dict = dict(sorted(dic.items(), reverse = True, key = lambda x:x[1]))
print(result_dict)
st_cnt = 1
dict_keys = list(result_dict.keys())
score = list(result_dict.values())
A_lst = list()
B_lst = list()
C_lst = list()
same_dic = dict()
same_Akeys = list()
same_Bkeys = list()
same_Ckeys = list()
i_same = 0
i = 0
A_same = list()
while(st_cnt < max_st - 1):
    now_st = 0
    if result_dict[dict_keys[i]] in score[i+1:]:
        i_same = score.count(result_dict[dict_keys[i]])
        if st_cnt - 1 + i_same <= (max_st - 1) * 0.3:
            now_st = st_cnt
            for k in range(now_st, now_st + i_same):
                A_lst.append(dict_keys[k - 1])
                same_Akeys.append(dict_keys[k - 1])
                if k+1 != now_st + i_same:
                    st_cnt += 1
            same_dic[score[i]] = i_same
        elif (max_st - 1) * 0.3 < st_cnt - 1 + i_same <= (max_st - 1) * 0.7:
            now_st = st_cnt
            for k in range(now_st, now_st + i_same):
                B_lst.append(dict_keys[k - 1])
                same_Bkeys.append(dict_keys[k - 1])
                if k + 1 != now_st + i_same:
                    st_cnt += 1
            same_dic[score[i]] = i_same
        else:
            now_st = st_cnt
            for k in range(now_st, now_st + i_same):
                C_lst.append(dict_keys[k - 1])
                same_Ckeys.append(dict_keys[k - 1])
                if k + 1 != now_st + i_same:
                    st_cnt += 1
            same_dic[score[i]] = i_same

    else:
        if st_cnt <= (max_st - 1) * 0.3:
            A_lst.append(dict_keys[i])
        elif st_cnt <= (max_st - 1) * 0.7:
            B_lst.append(dict_keys[i])
        else:
            C_lst.append(dict_keys[i])
    i = st_cnt
    st_cnt += 1

A_st = list()
B_st = list()
C_st = list()

for i in A_lst:
    if i not in A_st:
        A_st.append(i)
for i in B_lst:
    if i not in B_st:
        B_st.append(i)
for i in C_lst:
    if i not in C_st:
        C_st.append(i)
st_cnt = 1
i = 0
while(st_cnt <= len(A_st)):
    if A_st[i] in same_Akeys:
        s_cnt = score.count(result_dict[A_st[i]])
        if s_cnt + st_cnt - 1<= len(A_st) * 0.5:
            for a in range(s_cnt):
                ws.cell(row = A_st[i], column = 8, value = "A+")
                if a + 1 != s_cnt:
                    st_cnt += 1
                    i += 1
        else:
            for a in range(s_cnt):
                ws.cell(row = A_st[i], column = 8, value = "A0")
                if a+1 != s_cnt:
                    st_cnt += 1 
                    i += 1
    else:
        if st_cnt <= len(A_st) * 0.5:
            ws.cell(row = A_st[i], column = 8, value = "A+")
        else:
            ws.cell(row = A_st[i], column = 8, value = "A0")
    i = st_cnt
    st_cnt += 1
st_cnt = 1
i = 0
while(st_cnt <= len(B_st)):
    if B_st[i] in same_Bkeys:
        s_cnt = score.count(result_dict[B_st[i]])
        if s_cnt + st_cnt - 1 <= len(B_st) * 0.5:
            for a in range(s_cnt):
                ws.cell(row = B_st[i], column = 8, value = "B+")
                if a + 1 != s_cnt:
                    st_cnt += 1
                    i += 1
        else:
            for a in range(s_cnt):
                ws.cell(row = B_st[i], column = 8, value = "B0")
                if a + 1 != s_cnt:
                    st_cnt += 1
                    i += 1
    else:
        if st_cnt <= len(B_st) * 0.5:
            ws.cell(row = B_st[i], column = 8, value = "B+")
        else:
            ws.cell(row = B_st[i], column = 8, value = "B0")
    i = st_cnt
    st_cnt += 1
st_cnt = 1
i = 0
while(st_cnt <= len(C_st)):
    if C_st[i] in same_Ckeys:
        s_cnt = score.count(result_dict[C_st[i]])
        if s_cnt + st_cnt - 1 <= len(C_st) * 0.5:
            for a in range(s_cnt):
                ws.cell(row = C_st[i], column = 8, value = "C+")
                if a + 1 != s_cnt:
                    st_cnt += 1
                    i += 1
        else:
            for a in range(s_cnt):
                ws.cell(row = C_st[i], column = 8, value = "C0")
                if a + 1 != s_cnt:
                    st_cnt += 1
                    i += 1
    else:
        if st_cnt <= len(C_st) * 0.5:
            ws.cell(row = C_st[i], column = 8, value = "C+")
        else:
            ws.cell(row = C_st[i], column = 8, value = "C0")
    i = st_cnt
    st_cnt += 1

wb.save(filename = dest_filename)
